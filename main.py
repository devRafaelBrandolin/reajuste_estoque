#-------------------------------------
#------IMPORTANDO BIBLIOTECAS---------
#-------------------------------------
import tkinter as tk
from tkinter import ttk, messagebox
import subprocess
import os
import openpyxl
import datetime
from time import sleep

#-------------------------------------
#---------DEFININDO VARIAVEIS-----------
#-------------------------------------
nome_arquivo1 = 'rhp.xlsx'
nome_arquivo2 = 'silvestre.xlsx'
nome_arquivo3 = 'real.xlsx'
version = '2.0'

#-------------------------------------
#---------DEFININDO FUNÇÕES-----------
#-------------------------------------

prog = 0
def bar(valor):
    progressbar['value'] = (prog + valor) * 10
    janela.update()

def bar_complete():
    label_complete.place(relx=0.05, rely=0.72)
    # Atualiza a barra de progresso
    def update_progress(value):
        progressbar['value'] = value
        janela.update()

    # Cria uma animação de progresso completa
    for i in range(3):
        update_progress((prog + 0) * 10)
        sleep(0.1)
        update_progress((prog + 10) * 10)
        sleep(0.1)

    # Exibe a conclusão
    sleep(2)
    label_complete.place_forget()
    progressbar['value'] = 0
    janela.update()

def button_refresh():
    if os.path.isfile(nome_arquivo1) and os.path.isfile(nome_arquivo2):
        print(f'O arquivo existe na pasta do projeto.')
    else:
        messagebox.showerror("Error", f"O arquivo {nome_arquivo1} e ou {nome_arquivo2} não existe!!!\nExporte os relatórios do Wsac para continuar.")

def coluna_para_indice(coluna):
    return ord(coluna.upper()) - ord('A')

def entrada():
    #VERIFICANDO SE OS ARQUIVOS EXISTEM
    if os.path.isfile(nome_arquivo1) and os.path.isfile(nome_arquivo2):
        bar(1)
        #ABRINDO UMA PLANILHA DO EXCEL
        planilha = openpyxl.load_workbook(nome_arquivo1)
        #SELECIONANDO A PRIMEIRA ABA DA PLANILHA
        aba = planilha.active
        #PERCORRENDO TODAS AS LINHAS DA PLANILHA, COMEÇANDO PELA SEGUNDA LINHA
        for row in aba.iter_rows(min_row=2):
            #OBTENDO O VALOR DA CÉLULA NA COLUNA 2
            valor = row[1].value
            #VERIFICA SE O VALOR É MENOR DO QUE ZERO
            if valor < 0:
                #DEFINE O VALOR PARA 0
                row[1].value = 0
        #DELETA A PLANILHA ANTIGA
        os.remove(nome_arquivo1)
        bar(2)
        #SALVA A PLANILHA ATUALIZADA E CRIA UMA NOVA PLANILHA PARA USAR DE BASE
        planilha.save(nome_arquivo1)
        planilha.save(nome_arquivo3)
        
        #zREPETINDO OS PASSOS COM A PLANILHA 2
        planilha = openpyxl.load_workbook(nome_arquivo2)
        aba = planilha.active
        for row in aba.iter_rows(min_row=2):
            valor = row[2].value
            if valor < 0:
                row[2].value = 0
        os.remove(nome_arquivo2)
        bar(3)
        planilha.save(nome_arquivo2)
#----------------------------------------------------------------------------------
#---PLANILHAS ZERADAS, VAMOS IGUALAR OS VALORES------------------------------------
#----------------------------------------------------------------------------------
        #LENDO OS DADOS DAS PLANILHAS
        real_workbook = openpyxl.load_workbook(nome_arquivo3)
        silvestre_workbook = openpyxl.load_workbook(nome_arquivo2)

        #CRIANDO UMA NOVA PLANILHA - QUE RECEBERA OS DADOS 
        nova_planilha = openpyxl.Workbook()
        nova_planilha_ativa = nova_planilha.active

        #OBTENDO AS FOLHAS ATIVAS
        estoque_real_sheet = real_workbook.active
        silvestre_sheet = silvestre_workbook.active

        #OBTENDO O NUMERO MAXIMO DE LINHAS
        num_linhas_silvestre = silvestre_sheet.max_row
        bar(4)

        #PERCORRENDO CADA LINHA DA PLANILHA
        for linha_silvestre in range(2, num_linhas_silvestre + 1):
            codigo_silvestre = silvestre_sheet.cell(row=linha_silvestre, column=coluna_para_indice('B') + 1).value
            #VERIFICA SE O CAMPO NÃO ESTA VAZIO 
            if codigo_silvestre is not None:

                #PROCURANDO O CÓDIGO NA COLUNA A
                for linha_estoque_real in range(1, estoque_real_sheet.max_row + 1):
                    codigo_estoque_real = estoque_real_sheet.cell(row=linha_estoque_real, column=coluna_para_indice('A') + 1).value

                    #VERIFICA SE O CODIGO DA PLANILHA É IGUAL O DA OUTRA
                    if codigo_silvestre == codigo_estoque_real:
                        #OBTENDO OS VALORES DA CELULA A DIREITA DOS CODIGOS "CODIGO | VALOR"
                        valor_estoque_real = estoque_real_sheet.cell(row=linha_estoque_real, column=coluna_para_indice('B') + 1).value
                        valor_cod_silvestre = silvestre_sheet.cell(row=linha_silvestre, column=coluna_para_indice('A') + 1).value

                        #PREENCHENDO A NOVA PLANILHA COM ESSES VALORES
                        nova_planilha_ativa.append([valor_cod_silvestre, valor_estoque_real])

        #DELETA A PLANILHA ANTIGA
        os.remove(nome_arquivo2)
        bar(5)
        #SALVA A PLANILHA NOVA
        nova_planilha.save(nome_arquivo2)
#----------------------------------------------------------------------------------
#---CRIANDO ARQUIVO PARA IMPORTAR OS NOVOS DADOS-----------------------------------
#----------------------------------------------------------------------------------
        #CARREGANDO A PLANILHA
        workbook1 = openpyxl.load_workbook(nome_arquivo2)
        bar(6)

        #SELECIONA PLANILHA
        sheet1 = workbook1.active
        #OBTENDO DATA ATUAL
        data_atual = datetime.datetime.now().strftime("%d-%m-%Y")
        bar(7)

        #ABRE O ARQUIVO DE TEXTO PARA COLOCARMOS OS DADOS
        with open(f'estoque_silvestre {data_atual}.txt', 'w') as txt_file2:
            
            #ESCREVE OS CABEÇALHOS DAS COLUNAS NO ARQUIVO DE TEXTO
            headers2 = [coluna.value for coluna in sheet1[1]]
            txt_file2.write(';'.join(map(str, headers2)) + '\n')
            bar(8)

            #ITERA SOBRE AS LINHAS DA PLANILHA (começando da segunda linha, já que a primeira contém os cabeçalhos)
            for row in sheet1.iter_rows(min_row=2, values_only=True):
                #ESCREVE OS VALORES DE CADA COLUNA NO ARQUIVO DE TEXTO, SEPARANDOS POR ";"
                txt_file2.write(';'.join(map(str, row)) + '\n')
        
        #DELETA AS PLANILHAS ANTIGAS
        bar(9)
        os.remove(nome_arquivo1)
        os.remove(nome_arquivo2)
        bar(10)
        bar_complete()
    else:
        messagebox.showerror("Error", f"O arquivo {nome_arquivo1} e ou {nome_arquivo2} não existe!!!\nExporte os relatórios do Wsac para continuar.")
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------

def saida():
    #VERIFICA SE OS ARQUIVOS EXISTEM
    if os.path.isfile(nome_arquivo1) and os.path.isfile(nome_arquivo2):
        bar(1)
        #ABRE A PLANILHA DO EXCEL
        planilha = openpyxl.load_workbook(nome_arquivo1)
        #SELECIONA A PRIMEIRA ABA
        aba = planilha.active
        #PERCORRE CADA LINHA DA PLANILHA COMEÇANDO PELA SEGUNDA LINHA
        for row in aba.iter_rows(min_row=2):
            #OBTEM O VALOR DA CÉLULA NA COLUNA 2
            valor = row[1].value
            #VERIFICA SE É MENOR QUE 0
            if valor < 0:
                #DEFINE O VALOR PARA 0
                row[1].value = 0
        #DELETA A PLANILHA ANTIGA
        os.remove(nome_arquivo1)
        #SALVA A PLANILHA
        planilha.save(nome_arquivo1)
        bar(2)
       
        #REPETINDO O PROCESSO PARA A OUTRA PLANILHA
        planilha = openpyxl.load_workbook(nome_arquivo2)
        aba = planilha.active
        for row in aba.iter_rows(min_row=2):
            valor = row[2].value
            if valor < 0:
                row[2].value = 0
        os.remove(nome_arquivo2)
        planilha.save(nome_arquivo2)
        bar(3)

#----------------------------------------------------------------------------------
#---Planilhas zeradas vamos criar o arquivo de saída do estoque -------------------
#----------------------------------------------------------------------------------
        # Leitura das planilhas
        rhp_workbook = openpyxl.load_workbook(nome_arquivo1)
        silvestre_workbook = openpyxl.load_workbook(nome_arquivo2)
        estoque_real_workbook = openpyxl.load_workbook(nome_arquivo3)

        # Criando nova planilha
        nova_planilha = openpyxl.Workbook() #vamos salvar as saidas da silvestre aqui
        nova_planilha_ativa = nova_planilha.active
        nova_planilha2 = openpyxl.Workbook() #vamos salvar as saidas da rhp aqui
        nova_planilha2_ativa = nova_planilha2.active

        # Obter as folhas ativas
        estoque_real_sheet = estoque_real_workbook.active
        silvestre_sheet = silvestre_workbook.active
        rhp_sheet = rhp_workbook.active

        # Obtendo o número de linhas nas planilhas
        num_linhas_silvestre = silvestre_sheet.max_row

        # Para cada código na coluna B da planilha
        for linha_silvestre in range(2, num_linhas_silvestre + 1):
            codigo_silvestre = silvestre_sheet.cell(row=linha_silvestre, column=coluna_para_indice('B') + 1).value
            if codigo_silvestre is not None:

                # Procurar o código na coluna A da planilha estoque_real.xlsx
                for linha_estoque_real in range(1, estoque_real_sheet.max_row + 1):
                    codigo_estoque_real = estoque_real_sheet.cell(row=linha_estoque_real, column=coluna_para_indice('A') + 1).value

                    if codigo_silvestre == codigo_estoque_real:
                        # Obter os valores das células à direita
                        valor_estoque_real = estoque_real_sheet.cell(row=linha_estoque_real, column=coluna_para_indice('B') + 1).value
                        valor_silvestre = silvestre_sheet.cell(row=linha_silvestre, column=coluna_para_indice('C') + 1).value
                        valor_cod_silvestre = silvestre_sheet.cell(row=linha_silvestre, column=coluna_para_indice('A') + 1).value
                        valor_rhp = rhp_sheet.cell(row=linha_estoque_real, column=coluna_para_indice('B') + 1).value
                        
                        if valor_estoque_real > valor_silvestre:
                        # Calcular a diferença
                            diferenca = valor_estoque_real - valor_silvestre
                        else:
                            diferenca = 0
                        if valor_estoque_real > valor_rhp:
                            diferenca2 = valor_estoque_real - valor_rhp
                        else:
                            diferenca2 = 0
                        
                        saida_ = diferenca + diferenca2
                        saida_total = valor_estoque_real - saida_
                        print(saida_)
                        print(saida_total)

                        # Preencher a nova planilha
                        nova_planilha_ativa.append([valor_cod_silvestre, saida_total])
                        nova_planilha2_ativa.append([codigo_estoque_real, saida_total])
                    
        #deleta planilha antiga
        os.remove(nome_arquivo2)
        os.remove(nome_arquivo1)
        bar(4)
        # Salvar a nova planilha
        nova_planilha.save(nome_arquivo2)
        nova_planilha2.save(nome_arquivo1)      
        bar(5)

#----------------------------------------------------------------------------------
#---agora vamos criar o arquivo de importação do estoque para cada estoque---------
#----------------------------------------------------------------------------------

        # Verifica se a pasta 'backup' existe
        if not os.path.exists('backup'):
            # Se não existir, cria a pasta
            os.makedirs('backup')
            
        # Carrega a planilha
        workbook2 = openpyxl.load_workbook(nome_arquivo2)
        workbook1 = openpyxl.load_workbook(nome_arquivo1)

        # Seleciona a primeira planilha
        sheet2 = workbook2.active
        # Obtendo a data atual
        data_atual = datetime.datetime.now().strftime("%d-%m-%Y - %H-%M")
        bar(9)
        
        # Caminho completo para o arquivo de texto dentro da pasta 'backup'
        file_path = os.path.join('backup', f'estoque_silvestre {data_atual}.txt')


        # Abre o arquivo de texto para escrita
        with open(file_path, 'w') as txt_file2:
            
            # Escreve os cabeçalhos das colunas no arquivo de texto
            headers2 = [coluna.value for coluna in sheet2[1]]
            txt_file2.write(';'.join(map(str, headers2)) + '\n')

            # Itera sobre as linhas da planilha (começando da segunda linha, já que a primeira contém os cabeçalhos)
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                # Escreve os valores de cada coluna no arquivo de texto, separados por ponto e vírgula
                txt_file2.write(';'.join(map(str, row)) + '\n')

        # Seleciona a segunda planilha
        sheet2 = workbook1.active
        # Obtendo a data atual
        data_atual = datetime.datetime.now().strftime("%d-%m-%Y - %H-%M")

        # Caminho completo para o arquivo de texto dentro da pasta 'backup'
        file_path = os.path.join('backup', f'estoque_rhp {data_atual}.txt')


        # Abre o arquivo de texto para escrita
        with open(file_path, 'w') as txt_file2:
            
            # Escreve os cabeçalhos das colunas no arquivo de texto
            headers2 = [coluna.value for coluna in sheet2[1]]
            txt_file2.write(';'.join(map(str, headers2)) + '\n')

            # Itera sobre as linhas da planilha (começando da segunda linha, já que a primeira contém os cabeçalhos)
            for row in sheet2.iter_rows(min_row=2, values_only=True):
                # Escreve os valores de cada coluna no arquivo de texto, separados por ponto e vírgula
                txt_file2.write(';'.join(map(str, row)) + '\n')
                
        #deleta planilha antiga
        os.remove(nome_arquivo1)
        os.remove(nome_arquivo2)
        os.remove(nome_arquivo3)
        bar(10)
        bar_complete()        
    else:
        messagebox.showerror("Error", f"O arquivo {nome_arquivo1} e ou {nome_arquivo2} não existe!!!\nExporte os relatórios do Wsac para continuar.")

#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------
#-------------------------------------------------------------------------------------------

def button_entrada():
    entrada()
    
def button_saida():
    saida()

def sair():
    janela.destroy()

def sobre():
    tk.messagebox.showinfo("Sobre", "Esta aplicação foi desenvolvida para suprir a necessidade de manter atualizado o estoque da empresa.\nO objetivo é trabalhar com as planilhas de relatorio dos estoques e gerenciar os dados para importar um novo relatorio com o estoque atualizado!\nPara saber mais acesse o meu GitHub.")

def github():
    import webbrowser
    webbrowser.open("https://github.com/devRafaelBrandolin")

def abrir_pasta_projeto():
    try:
        subprocess.Popen(['explorer', '.'])  # No Windows
    except OSError:
        try:
            subprocess.Popen(['xdg-open', '.'])  # No Linux
        except OSError:
            tk.messagebox.showerror("Erro", "Não foi possível abrir a pasta do projeto.")

#-------------------------------------
#---------CRIANDO JANELA--------------
#-------------------------------------
janela = tk.Tk()
#DEFININDO UM TITULO PARA JANELA
janela.title(f"Reajuste - Wsac {version}")
#DEFININDO O TAMANHO DA JANELA
janela.geometry('300x220')
#IMPEDINDO QUE A JANELA SEJE REDIMENSIONÁVEL
janela.resizable(width=False, height=False)

# Caminho para o ícone (.ico no Windows)
icon_path = './box.ico'

# Definir o ícone da aplicação
janela.iconbitmap(icon_path)

#-------------------------------------
#-----CRIANDO WIDGETS DA JANELA-------
#-------------------------------------
#CRIANDO UM MENU PARA A JANELA
# Menu
barra_menu = tk.Menu(janela)
janela.config(menu=barra_menu)
# Menu "Arquivo"
menu_arquivo = tk.Menu(barra_menu, tearoff=0)
barra_menu.add_cascade(label="Arquivo", menu=menu_arquivo)
menu_arquivo.add_command(label="Abrir Pasta", command=abrir_pasta_projeto)
menu_arquivo.add_command(label="Sair", command=sair)
# Menu "Ajuda"
menu_ajuda = tk.Menu(barra_menu, tearoff=0)
barra_menu.add_cascade(label="Ajuda", menu=menu_ajuda)
menu_ajuda.add_command(label="Sobre", command=sobre)
menu_ajuda.add_command(label="Github", command=github)

#CRIANDO UM CABEÇALHO
#DEFININDO A COR DE FUNDO DO CABEÇALHO
cabecalho_frame = tk.Frame(janela, bg="#4CAF50")
#ESTENDENDO O CABEÇALHO À LARGURA DA JANELA
cabecalho_frame.grid(row=0, column=0, pady=10, columnspan=3, sticky="nsew")
#CRIANDO UM TITULO PARA O CABEÇALHO
label_criativo = tk.Label(cabecalho_frame, text="Bem-vindo ao REAJUSTE!", font=("Arial", 16, "bold"), fg="white", bg="#4CAF50")
label_criativo.grid(row=0, column=0, padx=10, pady=10)

#-------------------------------------
#------CRIANDO BOTÕES DA JANELA-------
#-------------------------------------
botao1 = tk.Button(janela, text="REFRESH", font=("Arial", 10, "bold"),  width='10', command=button_refresh)
botao1.grid(row=1, column=0, pady=5)

botao2 = tk.Button(janela, text="Entrada", font=("Arial", 10, "bold"), width='10', command=button_entrada)
botao2.grid(row=1, column=1, pady=5)

botao3 = tk.Button(janela, text="Saída", font=("Arial", 10, "bold"), width='10', command=button_saida)
botao3.grid(row=1, column=2, pady=5)

progressbar = ttk.Progressbar(janela, orient='horizontal', length=160, mode='determinate')
progressbar.place(relx=0.05, rely=0.61)

label_complete = tk.Label(janela,text='Complete!!!')
label_complete.place(relx=0.05, rely=0.72)
label_complete.place_forget()

botao6 = tk.Button(janela, text="SAIR", font=("Arial", 10, "bold"), width='10', command=sair)
botao6.grid(row=2, column=2, pady=8)

#-------------------------------------
# Rodapé
rodape_frame = tk.Frame(janela)
rodape_frame.grid(row=3, column=0, pady=10, columnspan=3)

rodape_label = tk.Label(rodape_frame, text="Desenvolvido por: Rafael Brandolin")
rodape_label.grid(row=0, column=0, pady=10)

janela.mainloop()